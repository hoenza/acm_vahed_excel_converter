import openpyxl
import xlsxwriter

ex_in = openpyxl.load_workbook('paye.xlsx')
sheet_in = ex_in['Sheet1']
ex_out = xlsxwriter.Workbook('output1.xlsx')
sheet_out = ex_out.add_worksheet('Merged')

for i in range(2, sheet_in.max_row + 1):
    p = str(sheet_in['A' + str(i)].value)
    p = p.replace('\n', '')
    sheet_out.write('A' + str(i + 1), p)
    p = str(sheet_in['B' + str(i)].value)
    p = p.replace('\n', '')
    sheet_out.write('B' + str(i + 1), p)
    p = str(sheet_in['C' + str(i)].value)
    p = p.replace('\n', '')
    sheet_out.write('C' + str(i + 1), p)
    p = str(sheet_in['D' + str(i)].value)
    p = p.replace('\n', '')
    sheet_out.write('D' + str(i + 1), p)
    p = str(sheet_in['E' + str(i)].value)
    p = p.replace('\n', '')
    sheet_out.write('E' + str(i + 1), p)
    p = str(sheet_in['N' + str(i)].value)
    p = p.replace('\n', '')
    sheet_out.write('P' + str(i + 1), p)
    quiz_time = str(sheet_in['L' + str(i)].value)
    if quiz_time != 'None':
        sheet_out.write('R' + str(i + 1), quiz_time[3:5])
        sheet_out.write('S' + str(i + 1), quiz_time[0:2])
        sheet_out.write('T' + str(i + 1), quiz_time[9:])
    p = str(sheet_in['F' + str(i + 1)].value)
    if p != 'None':
        j=0
        while j<len(p) and p[j] != '-':
            j+=1
        sheet_out.write('F' +str(i + 1), p[:j])
        sheet_out.write('G' + str(i + 1), p[j+1:])
    p = str(sheet_in['G' + str(i + 1)].value)
    if p != 'None':
        j = 0
        while j < len(p) and p[j] != '-':
            j += 1
        sheet_out.write('H' + str(i + 1), p[:j])
        sheet_out.write('I' + str(i + 1), p[j + 1:])
    p = str(sheet_in['H' + str(i + 1)].value)
    if p != 'None':
        j = 0
        while j < len(p) and p[j] != '-':
            j += 1
        sheet_out.write('J' + str(i + 1), p[:j])
        sheet_out.write('K' + str(i + 1), p[j + 1:])
    p = str(sheet_in['I' + str(i + 1)].value)
    if p != 'None':
        j = 0
        while j < len(p) and p[j] != '-':
            j += 1
        sheet_out.write('L' + str(i + 1), p[:j])
        sheet_out.write('M' + str(i + 1), p[j + 1:])
    p = str(sheet_in['J' + str(i + 1)].value)
    if p != 'None':
        j = 0
        while j < len(p) and p[j] != '-':
            j += 1
        sheet_out.write('N' + str(i + 1), p[:j])
        sheet_out.write('O' + str(i + 1), p[j + 1:])

    # tmp = 1
    # while i+tmp <= sheet_in.max_row and str(sheet_in['A' + str(i + tmp)].value) == 'None':
    #     tmp += 1
    # for k in range(0, tmp):
    #     tmp_time = sheet_in['O' + str(i + k)].value
    #     if str(tmp_time) == 'None':
    #         continue
    #     j = 0
    #     while j < len(tmp_time) and tmp_time[j] != ':':
    #         j += 1
    #     if tmp_time.startswith('امتحان'):
    #         sheet_out.write('R' + str(i-delay_cursor), tmp_time[12:14])
    #         sheet_out.write('S' + str(i-delay_cursor), tmp_time[15:17])
    #         sheet_out.write('T' + str(i-delay_cursor), tmp_time[26:37])
    #     else:
    #         tmp_time = tmp_time[j + 2:]
    #         if tmp_time.startswith('شنبه'):
    #             tmp_time = tmp_time[5:]
    #             sheet_out.write('F' + str(i + 1), tmp_time[:5])
    #             sheet_out.write('G' + str(i + 1), tmp_time[6:])
    #         elif tmp_time.startswith('يك شنبه'):
    #             tmp_time = tmp_time[8:]
    #             sheet_out.write('H' + str(i + 1), tmp_time[:5])
    #             sheet_out.write('I' + str(i + 1), tmp_time[6:])
    #         elif tmp_time.startswith('دو شنبه'):
    #             tmp_time = tmp_time[8:]
    #             sheet_out.write('J' + str(i + 1), tmp_time[:5])
    #             sheet_out.write('K' + str(i + 1), tmp_time[6:])
    #         elif tmp_time.startswith('سه شنبه'):
    #             tmp_time = tmp_time[8:]
    #             sheet_out.write('L' + str(i + 1), tmp_time[:5])
    #             sheet_out.write('M' + str(i + 1), tmp_time[6:])
    #         elif tmp_time.startswith('چهار شنبه'):
    #             tmp_time = tmp_time[10:]
    #             sheet_out.write('N' + str(i + 1), tmp_time[:5])
    #             sheet_out.write('O' + str(i + 1), tmp_time[6:])
print('ok')
