import openpyxl
import xlsxwriter
import sys
filename = sys.argv[1] if len(sys.argv) == 2 else 'input.xlsx'
ex_in = openpyxl.load_workbook(filename)
ex_out = xlsxwriter.Workbook(filename.split('.')[0]+'_output.xlsx')
sheet_out = ex_out.add_worksheet('Merged')
delay_cursor = -1
sheet_in = ex_in['Sheet1']
for i in range(2, sheet_in.max_row + 1):
    if str(sheet_in['A' + str(i)].value) != 'None':
        sheet_out.write('A' + str(i - delay_cursor), sheet_in['G' + str(i)].value)
        course_id_in = sheet_in['F' + str(i)].value
        # course_id_in = course_id_in.replace('_', '-')
        j = 0
        while j < len(course_id_in) and course_id_in[j] not in ['-', '_']:
            j += 1  
        sheet_out.write('B' + str(i - delay_cursor), course_id_in[:j])
        sheet_out.write('C' + str(i - delay_cursor), course_id_in[j + 1:])
        sheet_out.write('D' + str(i - delay_cursor), sheet_in['H' + str(i)].value)
        sheet_out.write('E' + str(i - delay_cursor), sheet_in['N' + str(i)].value)
        sheet_out.write('P' + str(i - delay_cursor), sheet_in['J' + str(i)].value)
        tmp = 1
        while i+tmp <= sheet_in.max_row and str(sheet_in['A' + str(i + tmp)].value) == 'None':
            tmp += 1
        for k in range(0, tmp):
            tmp_time = sheet_in['O' + str(i + k)].value
            if str(tmp_time) == 'None':
                continue
            j = 0
            while j < len(tmp_time) and tmp_time[j] != ':':
                j += 1
            if tmp_time.startswith('امتحان'):
                sheet_out.write('R' + str(i-delay_cursor), tmp_time[12:14])
                sheet_out.write('S' + str(i-delay_cursor), tmp_time[15:17])
                sheet_out.write('T' + str(i-delay_cursor), tmp_time[26:37])
            else:
                tmp_time = tmp_time[j + 2:]
                if tmp_time.startswith('شنبه'):
                    tmp_time = tmp_time[5:]
                    sheet_out.write('F' + str(i - delay_cursor), tmp_time[:5])
                    sheet_out.write('G' + str(i - delay_cursor), tmp_time[6:])
                elif tmp_time.startswith('يك شنبه'):
                    tmp_time = tmp_time[8:]
                    sheet_out.write('H' + str(i - delay_cursor), tmp_time[:5])
                    sheet_out.write('I' + str(i - delay_cursor), tmp_time[6:])
                elif tmp_time.startswith('دو شنبه'):
                    tmp_time = tmp_time[8:]
                    sheet_out.write('J' + str(i - delay_cursor), tmp_time[:5])
                    sheet_out.write('K' + str(i - delay_cursor), tmp_time[6:])
                elif tmp_time.startswith('سه شنبه'):
                    tmp_time = tmp_time[8:]
                    sheet_out.write('L' + str(i - delay_cursor), tmp_time[:5])
                    sheet_out.write('M' + str(i - delay_cursor), tmp_time[6:])
                elif tmp_time.startswith('چهار شنبه'):
                    tmp_time = tmp_time[10:]
                    sheet_out.write('N' + str(i - delay_cursor), tmp_time[:5])
                    sheet_out.write('O' + str(i - delay_cursor), tmp_time[6:])
    else:
        delay_cursor += 1
ex_in.close()
ex_out.close()
print('ok')
